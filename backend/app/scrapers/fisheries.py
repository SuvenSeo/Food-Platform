"""
Sri Lanka Fisheries weekly fish-price scraper.

The Ministry of Fisheries publishes weekly fish-price Excel/PDF files under:
https://www.fisheries.gov.lk/web/index.php/en/statistics/weekly-fish-prices
"""

from __future__ import annotations

import io
import logging
import re
from calendar import monthrange
from datetime import datetime, timezone
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

FISHERIES_PRICE_URL = "https://www.fisheries.gov.lk/web/index.php/en/statistics/weekly-fish-prices"
FISHERIES_BASE_URL = "https://www.fisheries.gov.lk"

WEEK_NUMBERS = {"1st": 1, "2nd": 2, "3rd": 3, "4th": 4, "5th": 5}
MONTH_NUMBERS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

FISH_NAME_HINTS = (
    "Yellowfin tuna",
    "Skipjack tuna",
    "Indian Mackerel",
    "Trevally (L)",
    "Rock fish (L)",
    "Rock fish (S)",
    "Sail fish",
    "Sardinella",
    "Pony fish",
    "Katuwalla",
    "Anchovies",
    "Anchovy",
    "Prawns (M)",
    "Marlins",
    "Herrings",
    "Sharks",
    "Shark",
    "Seer (Ni-L)",
    "Seer",
)


def _parse_week_date(label: str) -> datetime:
    match = re.search(
        r"\b(1st|2nd|3rd|4th|5th)\s+Week(?:\s+of)?\s+([A-Za-z]+)\s+(\d{4})\b",
        label,
        flags=re.IGNORECASE,
    )
    if not match:
        return datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

    week = WEEK_NUMBERS[match.group(1).lower()]
    month = MONTH_NUMBERS[match.group(2).lower()]
    year = int(match.group(3))
    day = min(1 + ((week - 1) * 7), monthrange(year, month)[1])
    return datetime(year, month, day, tzinfo=timezone.utc)


def _price(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    cleaned = str(value).replace(",", "").replace(" ", "").strip()
    if not cleaned or cleaned in {"-", "nan"}:
        return None
    try:
        parsed = float(cleaned)
    except ValueError:
        return None
    return parsed if parsed > 0 else None


def _workbook_sheet_meta(sheet_name: str) -> tuple[str, str] | None:
    normalized = sheet_name.strip().lower()
    if "wholesale" in normalized:
        return ("Peliyagoda Fish Market (Fisheries)", "wholesale")
    if "retail" in normalized:
        return ("Selected Retail Markets (Fisheries)", "retail")
    return None


def parse_fisheries_workbook(content: bytes, *, label: str) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    quoted_at = _parse_week_date(label).isoformat()
    quotes: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        meta = _workbook_sheet_meta(sheet_name)
        if not meta:
            continue
        market_name, price_type = meta
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=4, values_only=True):
            common_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            current_price = _price(row[5] if len(row) > 5 else None)
            if not common_name or not current_price:
                continue
            quotes.append(
                {
                    "district": "Colombo" if price_type == "wholesale" else "Sri Lanka",
                    "market_name": market_name,
                    "item_name": common_name,
                    "category": "fish",
                    "unit": "kg",
                    "price_lkr": round(current_price, 2),
                    "source": "fisheries",
                    "quoted_at": quoted_at,
                    "notes": f"Fisheries weekly {price_type} fish price report: {label}.",
                }
            )
    return quotes


def _pdf_common_name(line: str) -> str | None:
    lowered = re.sub(r"\s+", " ", line).lower()
    compact = re.sub(r"[^a-z]+", "", lowered)
    for hint in FISH_NAME_HINTS:
        hint_lower = hint.lower()
        hint_compact = re.sub(r"[^a-z]+", "", hint_lower)
        if hint_lower in lowered or hint_compact in compact:
            return hint

    ascii_prefix = "".join(ch if ord(ch) < 128 else " " for ch in line)
    first_price = re.search(r"(?:\d{1,3}\s*,\s*)?\d{2,4}\.\d{2}", ascii_prefix)
    if not first_price:
        return None
    prefix = re.sub(r"^\s*\d+\s+", "", ascii_prefix[: first_price.start()])
    prefix = re.sub(r"[^A-Za-z0-9()/-]+", " ", prefix).strip(" -")
    prefix = re.sub(r"^[()\s]+", "", prefix).replace("R)ock", "Rock")
    return prefix or None


def parse_fisheries_pdf(content: bytes, *, label: str) -> list[dict[str, object]]:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("Fisheries: install pdfplumber to parse PDF reports")
        return []

    price_pattern = re.compile(r"(?:\d{1,3}\s*,\s*)?\d{2,4}\.\d{2}")
    quoted_at = _parse_week_date(label).isoformat()
    quotes: list[dict[str, object]] = []

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page_index, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            price_type = "wholesale" if page_index == 0 or "Wholesale" in text[:250] else "retail"
            market_name = (
                "Peliyagoda Fish Market (Fisheries)"
                if price_type == "wholesale"
                else "Selected Retail Markets (Fisheries)"
            )
            for line in text.splitlines():
                if not re.match(r"^\s*\d+\s+", line):
                    continue
                name = _pdf_common_name(line)
                prices = [_price(match.group(0)) for match in price_pattern.finditer(line)]
                prices = [price for price in prices if price is not None]
                if not name or len(prices) < 3:
                    continue
                quotes.append(
                    {
                        "district": "Colombo" if price_type == "wholesale" else "Sri Lanka",
                        "market_name": market_name,
                        "item_name": name,
                        "category": "fish",
                        "unit": "kg",
                        "price_lkr": round(float(prices[2]), 2),
                        "source": "fisheries",
                        "quoted_at": quoted_at,
                        "notes": f"Fisheries weekly {price_type} fish price report: {label}.",
                    }
                )
    return quotes


def _candidate_links(html: str) -> list[tuple[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    candidates: list[tuple[str, str]] = []
    for anchor in soup.find_all("a"):
        label = " ".join(anchor.get_text(" ", strip=True).split())
        href = anchor.get("href") or ""
        lower_href = href.lower()
        if "week" not in label.lower():
            continue
        if not lower_href.endswith((".xlsx", ".xls", ".pdf")):
            continue
        candidates.append((label, urljoin(FISHERIES_BASE_URL, href)))
    return candidates


def fetch_fisheries_market_quotes(timeout: float = 30.0, max_files: int = 4) -> list[dict[str, object]]:
    headers = {"User-Agent": "FoodLKBot/1.0 (+https://food-platform-one.vercel.app)"}
    with httpx.Client(follow_redirects=True, timeout=timeout, headers=headers) as client:
        index_response = client.get(FISHERIES_PRICE_URL)
        index_response.raise_for_status()
        for label, url in _candidate_links(index_response.text)[:max_files]:
            try:
                response = client.get(url)
                response.raise_for_status()
                if url.lower().endswith(".pdf"):
                    quotes = parse_fisheries_pdf(response.content, label=label)
                else:
                    quotes = parse_fisheries_workbook(response.content, label=label)
                if quotes:
                    logger.info("Fisheries: parsed %d quotes from %s", len(quotes), url)
                    return quotes
            except Exception as exc:
                logger.warning("Fisheries: failed to parse %s: %s", url, exc)
    logger.warning("Fisheries: no weekly fish price rows parsed")
    return []
